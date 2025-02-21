import pandas as pd
import requests
import tkinter as tk
import numpy as np
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
#botão para import
def load_local_file(file_path):
    return pd.read_excel(file_path, index_col=None)
def importar_planilha():
    global planilha_original  
    if 'planilha_original' not in globals():  
        planilha_path = filedialog.askopenfilename(title="Selecionar a Planilha Original", filetypes=[("Excel files", "*.xlsx")])
        if planilha_path:
            planilha_original = pd.read_excel(planilha_path)  # Carrega o arquivo Excel
            print(f"Planilha '{planilha_path}' importada com sucesso.")
        else:
            print("Nenhum arquivo selecionado.")
    else:
        print("A planilha já foi importada anteriormente.")
root = tk.Tk()
root.title("Importar Planilha")
btn_importar = tk.Button(root, text="Importar Planilha Original", command=importar_planilha)
btn_importar.pack(pady=20)
root.mainloop()
# Carregar as bases Excel
base_aeroportos = pd.read_excel('Base_Aeroportos.xlsx')
base_classe = pd.read_excel('Base_Classe.xlsx')

# Converter as bases para JSON
base_aeroportos.to_json('Base_Aeroportos.json', orient="records")
base_classe.to_json('Base_Classe.json', orient="records")

# Carregar as bases JSON como DataFrames
base_aeroportos = pd.read_json('Base_Aeroportos.json', orient="records")
base_classe = pd.read_json('Base_Classe.json', orient="records")

#certificar carregamento
if 'planilha_original' in globals():
    print(planilha_original.head())  
else:
    print("A planilha não foi importada. Por favor, tente novamente.")
# Certificar-se de que as colunas de data estão no formato correto
planilha_original['Data Emissão'] = pd.to_datetime(planilha_original['Data Emissão'], dayfirst=True, errors='coerce')
planilha_original['Data IN'] = pd.to_datetime(planilha_original['Data IN'], dayfirst=True, errors='coerce')
planilha_original['Data OUT'] = pd.to_datetime(planilha_original['Data OUT'], dayfirst=True, errors='coerce')
# Função para buscar o câmbio na API de acordo com a data
def buscar_cambio(url):
    # Faz a requisição à API
    response = requests.get(url)
    if response.status_code != 200:
        raise Exception("Falha ao obter os dados da API")
    # Extrai os dados da resposta JSON
    data = response.json().get('value', [])
    # Cria uma tabela com as datas e os valores do dólar
    df_cambio = pd.DataFrame(data)
    df_cambio['Data'] = pd.to_datetime(df_cambio['dataHoraCotacao']).dt.date  # Extrai apenas a data
    df_cambio.rename(columns={'cotacaoCompra': 'Câmbio'}, inplace=True)
    df_cambio = df_cambio[['Data', 'Câmbio']]  # Reordena as colunas
    return df_cambio
# Chamada da função com o link fornecido
url = "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)?@dataInicial='01-01-2023'&@dataFinalCotacao='10-01-2035'&$top=10000&$format=json&$select=cotacaoCompra,dataHoraCotacao"
df_cambio = buscar_cambio(url)

# Função para buscar cidade do aeroporto
def buscar_municipio(iata_code):
    resultado = base_aeroportos.loc[base_aeroportos['iata_code'] == iata_code, 'municipality']
    if not resultado.empty:
        return resultado.iloc[0]
    return None

# Função para busca da descrição de classe de acordo com nosso depara
def buscar_descricao_classe(cod_classe, cia_aereo):
    resultado = base_classe.loc[
        (base_classe['SIGLA'] == cod_classe) & (base_classe['CIA'] == cia_aereo),
        'CLASSE'
    ]
    if not resultado.empty:
        return resultado.iloc[0]
    return None

#Função para fazer o tratamento das colunas tarifarias
def ajustar_tarifas(df):
    df['Tarifa Cheia R$'] = df.apply(
        lambda x: x['Tarifa Paga R$']
        if pd.isna(x['Tarifa Cheia R$']) or x['Tarifa Cheia R$'] == "" or x['Tarifa Cheia R$'] < x['Tarifa Paga R$']
        else x['Tarifa Cheia R$'],
        axis=1
    )
    df['Tarifa Sugerida R$'] = df.apply(
        lambda x: x['Tarifa Paga R$'] if x['Tarifa Sugerida R$'] == 0 else x['Tarifa Sugerida R$'],
        axis=1
    )
    df['Tarifa Cheia R$'] = df.apply(
        lambda x: max(x['Tarifa Cheia R$'], x['Tarifa Mercado R$'], x['Tarifa Sugerida R$'], x['Tarifa Paga R$']),
        axis=1
    )
    df['Tarifa Sugerida R$'] = df.apply(
        lambda x: min(x['Tarifa Paga R$'], x['Tarifa Sugerida R$']),
        axis=1
    )
    return df

#Função para zerar as tarifas caso reemissão seja 'S'
def zerar_tarifas_economias(df):
    df.loc[df['Reemissão'] == 'S', [
        'Tarifa Cheia R$', 'Tarifa Mercado R$', 'Tarifa Sugerida R$', 'Tarifa Paga R$', 'Tarifa de Embarque R$',
        'Total (Tarifa + Taxa) R$', 'Saving(R$)', 'Eco. Obtida (R$)', 'Eco. Obtida (%)', 'Economia Não Obtida (R$)', 'Eco. Não Obtida (%)'
    ]] = 0

#Mapeamento das Cias
mapeamento_cias = {
    'LATAM': 'Latam',
    'GOL LINHAS AEREAS RIO DE JANEIRO': 'Gol',
    'AZUL LINHAS AEREAS BRASILEIRAS': 'Azul',
    'PACIFIC COASTAL AIRLINES': 'Pacific',
    'AVIANCA': 'Avianca',
    'COPA AIRLINES': 'Copa Airlines',
    'AEROLINEAS ARGENTINAS': 'Aerolineas',
    'AMERICAN AIRLINES': 'American Airlines',
    'AIR CANADA': 'Air Canada',
    'UNITED AIRLINES': 'United Airlines',
    'LATAM AIRLINES': 'Latam',
    'GOL LINHAS AÉREAS': 'Gol',
    'AIR FRANCE': 'Air France',
    'SWISS': 'Swiss',
    'AIR EUROPA LINHAS AEREAS': 'Air Europa',
    'KLM': 'KLM',
    'IBERIA': 'Iberia',
    'ALITALIA': 'Alitalia',
    'DELTA AIR LINES INC': 'Delta',
    'BRITISH AIRWAYS': 'British Airways',
    'QATAR AIRWAYS': 'Qatar Airways',
    'AEROMEXICO': 'Aeromexico',
    'TURKISH AIRLINES': 'Turkish Airlines',
    'TAP': 'TAP',
    'EMIRATES AIRLINES': 'Emirates',
    'LUFTHANSA': 'Lufthansa',
    'SCANDINAVIAN AIRLINES': 'Scandinavian Airlines',
    'GOL LINHAS AEREAS - G3': 'Gol'
}

#Criar mine banco para extração de Reason Codes
def criar_mini_banco(df):
    mini_banco = {}
    for item in df['Info. Referencial 1'].dropna():
        if '-' in item:
            cod, descricao = item.split('-', 1)
            mini_banco[descricao.strip().upper()] = cod.strip().upper()
    return mini_banco
#Função para extrair do mini_banco sigla e descrição do reason code
def buscar_codigo_motivo(descricao, mini_banco):
    return mini_banco.get(descricao, '')
mini_banco = criar_mini_banco(planilha_original)

#Gerar as sheets gerencial_aereo e gerencial_terrestre
def gerar_bases_gerenciais(df):
    gerencial_aereo = pd.DataFrame()
    gerencial_terrestre = pd.DataFrame()
    #DESENVOLVENDO A ABA GERENCIAL_AEREO
    df_aereo = df[df['Movimento'].str.contains("Aéreo", case=False, na=False)]
    if not df_aereo.empty:
        gerencial_aereo['Nome Cliente'] = df_aereo['Cliente']
        gerencial_aereo['Nome do passageiro'] = df_aereo['Passageiro']
        gerencial_aereo['Localizador'] = df_aereo['Localizador']
        gerencial_aereo['Bilhete'] = df_aereo['Bilhete']
        gerencial_aereo['Data Emissão'] = df_aereo.apply(
            lambda row: row['Data IN'] if row['Data Emissão'] > row['Data IN'] else row['Data Emissão'],
            axis=1
        )
        gerencial_aereo['Mês'] = gerencial_aereo['Data Emissão'].dt.strftime('%b')
        posicao_nome_cliente = gerencial_aereo.columns.get_loc('Nome Cliente')
        gerencial_aereo.insert(posicao_nome_cliente, 'Mês', gerencial_aereo.pop('Mês'))
        gerencial_aereo['Data Embarque'] = df_aereo['Data IN']
        gerencial_aereo['Data Retorno'] = df_aereo['Data OUT']
        gerencial_aereo['Antecedencia'] = (df_aereo['Data IN'] - gerencial_aereo['Data Emissão']).dt.days
        gerencial_aereo['CIA Aereo'] = df_aereo['Fornecedor'].apply(
            lambda x: mapeamento_cias.get(x, x)
            )
        gerencial_aereo['Cod Classe'] = df_aereo['Classe'].str[0]
        gerencial_aereo['Tipo Viagem(N/I)'] = df_aereo['Abrangência']
        gerencial_aereo['Descrição da classe'] = gerencial_aereo.apply(
            lambda x: buscar_descricao_classe(x['Cod Classe'], x['CIA Aereo']) 
                    if buscar_descricao_classe(x['Cod Classe'], x['CIA Aereo']) 
                    else ('Econômica' if x['Tipo Viagem(N/I)'] == 'Nacional' else 'Indefinido'),
            axis=1
        )
        # Reposiciona a coluna 'Descrição da classe' logo após 'Cod Classe'
        cols = list(gerencial_aereo.columns)
        idx = cols.index('Cod Classe')  # Encontra o índice de 'Cod Classe'
        cols.insert(idx + 1, cols.pop(cols.index('Descrição da classe')))  # Move a coluna
        gerencial_aereo = gerencial_aereo[cols]
        gerencial_aereo['Cidade Origem'] = df_aereo['Trechos'].str[:3].apply(buscar_municipio)
        gerencial_aereo['Cidade Destino'] = df_aereo['Destino Serviço'].apply(buscar_municipio)
        gerencial_aereo['Rota'] = df_aereo['Trechos']
        gerencial_aereo['Centro de custo'] = df_aereo['Centro de Custo']
        gerencial_aereo['Matrícula'] = df_aereo['Matrícula']
        gerencial_aereo['Requisição'] = df_aereo['OS']
        gerencial_aereo['Departamento'] = df_aereo['Divisão']
        gerencial_aereo['Solicitante'] = df_aereo['Solicitante']
        gerencial_aereo['Aprovador'] = df_aereo['Aprovador']
        gerencial_aereo['Projeto'] = df_aereo['Projeto']
        gerencial_aereo['Código de autorização'] = df_aereo['Requisição']
        gerencial_aereo['Motivo'] = df_aereo['Motivo']
        gerencial_aereo['Descricao da Viagem'] = df_aereo['Nível Funcionário']
        gerencial_aereo['Empenho'] = df_aereo['Empenho']
        gerencial_aereo['Informações Extras'] = df_aereo['Informações Extras']
        gerencial_aereo['Informação de Controle'] = df_aereo['Finalidade']
        gerencial_aereo['Tipo Pedido'] = df_aereo['Origem pedido'].str[:2]
        gerencial_aereo['Tipo requisição'] = df_aereo['Origem pedido'].str[:2] + df_aereo['OS'].astype(str)
        gerencial_aereo['Adesão'] = df_aereo.apply(
            lambda x: "ONLINE" if isinstance(x['Origem pedido'], str) and any(substring in x['Origem pedido'] for substring in ["SS", "SN"]) else "OFFLINE",
            axis=1
        )
        gerencial_aereo['Tipo Pagamento'] = df_aereo['Forma de Pagamento']
        # Converte as colunas de interesse do df_aereo para float com verificação de erros
        gerencial_aereo['Tarifa Cheia R$'] = pd.to_numeric(df_aereo.get('Tarifa Máxima', 0), errors='coerce')
        gerencial_aereo['Tarifa Mercado R$'] = pd.to_numeric(df_aereo.get('Total', 0), errors='coerce')
        gerencial_aereo['Tarifa Sugerida R$'] = pd.to_numeric(df_aereo.get('Tarifa Mínima', 0), errors='coerce')
        gerencial_aereo['Tarifa Paga R$'] = pd.to_numeric(df_aereo.get('Tarifa', 0), errors='coerce')
        gerencial_aereo['Tarifa de Embarque R$'] = pd.to_numeric(df_aereo.get('Taxas', 0), errors='coerce')

        # Calcula a coluna 'Total (Tarifa + Taxa) R$' garantindo que as somas sejam consistentes
        gerencial_aereo['Total (Tarifa + Taxa) R$'] = (
            gerencial_aereo['Tarifa Paga R$'].fillna(0) + gerencial_aereo['Tarifa de Embarque R$'].fillna(0)
        )
        gerencial_aereo['Reemissão'] = df_aereo['Reemissão']

        ajustar_tarifas(gerencial_aereo)
        zerar_tarifas_economias(gerencial_aereo)
       
        # Primeiro, ajuste a coluna 'Descrição Reason Code' com as condições especificadas
        gerencial_aereo['Descrição Reason Code'] = gerencial_aereo.apply(
            lambda row: (
                "REEMISSÃO" if row['Reemissão'] == 'S' else  
                "MENOR TARIFA ACEITA" if row['Tarifa Sugerida R$'] == row['Tarifa Paga R$'] else
                "PREFERENCIAL" if row['Tarifa Sugerida R$'] < row['Tarifa Paga R$'] else
                ''
            ),
            axis=1
        )
        # Agora, mapeie os valores de 'Descrição Reason Code' para os códigos especificados
        gerencial_aereo['Código Reason Code'] = gerencial_aereo['Descrição Reason Code'].map({
            "PREFERENCIAL": "CA",
            "MENOR TARIFA ACEITA": "LA",
            "REEMISSÃO": "EX"
        })
        colunas = list(gerencial_aereo.columns)
        # Posicionar 'Código Reason Code' antes da 'Rota'
        if 'Rota' in colunas and 'Código Reason Code' in colunas:
            rota_idx = colunas.index('Rota')
            colunas.insert(rota_idx + 1, colunas.pop(colunas.index('Código Reason Code')))
            gerencial_aereo = gerencial_aereo[colunas]
        # Posicionar 'Descrição Reason Code' antes de 'Centro de custo'
        if 'Centro de custo' in colunas and 'Descrição Reason Code' in colunas:
            centro_custo_idx = colunas.index('Centro de custo')
            colunas.insert(centro_custo_idx, colunas.pop(colunas.index('Descrição Reason Code')))
            gerencial_aereo = gerencial_aereo[colunas]
        gerencial_aereo['Saving(R$)'] = gerencial_aereo['Tarifa Mercado R$'] - gerencial_aereo['Tarifa Paga R$']
        gerencial_aereo['Eco. Obtida (R$)'] = gerencial_aereo['Tarifa Cheia R$'] - gerencial_aereo['Tarifa Paga R$']
        gerencial_aereo['Eco. Obtida (%)'] = np.where(
            gerencial_aereo['Tarifa Cheia R$'] == 0,
            0,
            gerencial_aereo['Eco. Obtida (R$)'] / gerencial_aereo['Tarifa Cheia R$']
        )
        gerencial_aereo['Economia Não Obtida (R$)'] = gerencial_aereo['Tarifa Sugerida R$'] - gerencial_aereo['Tarifa Paga R$']
        gerencial_aereo['Eco. Não Obtida (%)'] = np.where(
            gerencial_aereo['Tarifa Sugerida R$'] == 0,
            0,
            gerencial_aereo['Economia Não Obtida (R$)'] / gerencial_aereo['Tarifa Sugerida R$']
        )
        # Coluna Cambio
        gerencial_aereo['Data Emissão'] = pd.to_datetime(gerencial_aereo['Data Emissão']).dt.date
        df_cambio['Data'] = pd.to_datetime(df_cambio['Data']).dt.date
        # Realizando o merge para incluir a coluna "Câmbio" com base na data
        gerencial_aereo = gerencial_aereo.merge(df_cambio, left_on='Data Emissão', right_on='Data', how='left')
        gerencial_aereo.drop(columns=['Data'], inplace=True)  # Remove coluna duplicada "Data" após o merge
        # Posicionar a coluna "Câmbio" antes da 'Tarifa Cheia R$'
        posicao_tarifa = gerencial_aereo.columns.get_loc('Tarifa Cheia R$')
        gerencial_aereo.insert(posicao_tarifa, 'Câmbio', gerencial_aereo.pop('Câmbio'))
        #Função para criar a linha total no final do relatório
        total_aereo = pd.DataFrame({
            'Tarifa Cheia R$': [gerencial_aereo['Tarifa Cheia R$'].sum()],
            'Tarifa Mercado R$': [gerencial_aereo['Tarifa Mercado R$'].sum()],
            'Tarifa Sugerida R$': [gerencial_aereo['Tarifa Sugerida R$'].sum()],
            'Tarifa Paga R$': [gerencial_aereo['Tarifa Paga R$'].sum()],
            'Tarifa de Embarque R$': [gerencial_aereo['Tarifa de Embarque R$'].sum()],
            'Total (Tarifa + Taxa) R$': [gerencial_aereo['Total (Tarifa + Taxa) R$'].sum()],
            'Saving(R$)': [gerencial_aereo['Saving(R$)'].sum()],
            'Eco. Obtida (R$)': [gerencial_aereo['Eco. Obtida (R$)'].sum()],
            'Eco. Obtida (%)': [gerencial_aereo['Eco. Obtida (R$)'].sum() / gerencial_aereo['Tarifa Cheia R$'].sum()],
            'Economia Não Obtida (R$)': [gerencial_aereo['Economia Não Obtida (R$)'].sum()],
            'Eco. Não Obtida (%)': [gerencial_aereo['Economia Não Obtida (R$)'].sum() / gerencial_aereo['Tarifa Sugerida R$'].sum()]
        }, index=['Total'])
        gerencial_aereo = pd.concat([gerencial_aereo, total_aereo])
    
    #DESENVOLVENDO A ABA GERENCIAL_TERRESTREE
    df_terrestre = df[~df['Movimento'].str.contains(r'\bA[eé]reo\b', case=False, na=False)]
    if not df_terrestre.empty:
        gerencial_terrestre['Mês Mov'] = df_terrestre['Data Emissão'].dt.strftime('%b')
        gerencial_terrestre['Nome Cliente'] = df_terrestre['Cliente']
        gerencial_terrestre['Nome do passageiro'] = df_terrestre['Passageiro']
        gerencial_terrestre['Solicitante'] = df_terrestre['Solicitante']
        gerencial_terrestre['Localizador'] = df_terrestre['Localizador']
        gerencial_terrestre['Data Emissão'] = df_terrestre.apply(
            lambda row: row['Data IN'] if row['Data Emissão'] > row['Data IN'] else row['Data Emissão'],
            axis=1
        )
        gerencial_terrestre['Data Embarque'] = df_terrestre['Data IN']
        gerencial_terrestre['Data Retorno'] = df_terrestre['Data OUT']
        gerencial_terrestre['Antecedencia'] = (df_terrestre['Data IN'] - df_terrestre['Data Emissão']).dt.days
        gerencial_terrestre['Tipo Fornecedor'] = df_terrestre.apply(
            lambda x: "Locadora Nacional" if x['Abrangência'] == "Nacional" and x['Movimento'] == "CARRO" else (
                "Locadora Internacional" if x['Abrangência'] == "Internacional" and x['Movimento'] == "CARRO" else (
                "Hotel Nacional" if x['Abrangência'] == "Nacional" and x['Movimento'] == "HOTEL" else (
                "Hotel Internacional" if x['Abrangência'] == "Internacional" and x['Movimento'] == "HOTEL" else "Serviços diversos"))), axis=1
        )
        gerencial_terrestre['Nome Fornecedor'] = df_terrestre['Fornecedor']
        gerencial_terrestre['Municipio Fornecedor'] = df_terrestre['Município Fornecedor']
        gerencial_terrestre['Centro de Custo'] = df_terrestre['Centro de Custo']
        gerencial_terrestre['Matrícula'] = df_terrestre['Matrícula']
        gerencial_terrestre['Requisição'] = df_terrestre['OS']
        gerencial_terrestre['Departamento'] = df_terrestre['Divisão']
        gerencial_terrestre['Solicitante'] = df_terrestre['Solicitante']
        gerencial_terrestre['Aprovador'] =  df_terrestre['Aprovador']
        gerencial_terrestre['Projeto'] = df_terrestre['Projeto']
        gerencial_terrestre['Código de autorização'] = df_terrestre['Requisição']
        gerencial_terrestre['Motivo'] = df_terrestre['Motivo']
        gerencial_terrestre['Descricao da Viagem'] = df_terrestre['Nível Funcionário']
        gerencial_terrestre['Empenho'] = df_terrestre['Empenho']
        gerencial_terrestre['Informações Extras'] = df_terrestre['Informações Extras']
        gerencial_terrestre['Informação de Controle'] = df_terrestre['Finalidade']
        gerencial_terrestre['Tipo requisição'] = df_terrestre['Origem pedido'].str[:2] + df_terrestre['OS'].astype(str)       
        gerencial_terrestre['Tipo Pedido'] = df_terrestre['Origem pedido'].str[:2]
        gerencial_terrestre['Adesão'] = df_terrestre.apply(
            lambda x: "ONLINE" if isinstance(x['Origem pedido'], str) and any(substring in x['Origem pedido'] for substring in ["SS", "SN"]) else "OFFLINE",
            axis=1
        )
        gerencial_terrestre['Tipo Pagamento'] = df_terrestre['Forma de Pagamento']
        gerencial_terrestre['Reemissão'] = df_terrestre['Reemissão']
        # Atualiza o campo 'Motivo Solicitacao' sem comparação de tarifas e com verificação de tipo
        gerencial_terrestre['Descrição Reason Code'] = df_terrestre['Info. Referencial 1'].apply(
            lambda x: "Menor Tarifa Aceita" if pd.isna(x) or x == "" else "Preferencial" if isinstance(x, str) else ""
        )
        # Atribui o código do motivo usando a função buscar_codigo_motivo
        gerencial_terrestre['Sigla Reason Code'] = gerencial_terrestre['Descrição Reason Code'].apply(
            lambda motivo: "LA" if motivo == "Menor Tarifa Aceita" else "XX" if motivo == "Preferencial" else buscar_codigo_motivo(motivo, mini_banco)
        )
        col_municipio_idx = gerencial_terrestre.columns.get_loc('Municipio Fornecedor')
        sigla_reason_code = gerencial_terrestre.pop('Sigla Reason Code')
        gerencial_terrestre.insert(col_municipio_idx + 1, 'Sigla Reason Code', sigla_reason_code)
        # Move 'Descrição Reason Code' para antes de 'Centro de Custo'
        col_centro_custo_idx = gerencial_terrestre.columns.get_loc('Centro de Custo')
        descricao_reason_code = gerencial_terrestre.pop('Descrição Reason Code')
        gerencial_terrestre.insert(col_centro_custo_idx, 'Descrição Reason Code', descricao_reason_code)
        gerencial_terrestre['Diárias (Qted)'] = (df_terrestre['Data OUT'] - df_terrestre['Data IN']).dt.days.apply(lambda x: 1 if x == 0 else x)
        gerencial_terrestre['Tarifa Terrestre R$'] = pd.to_numeric(df_terrestre['Tarifa'], errors='coerce')
        gerencial_terrestre['Taxa Terrestre R$'] = pd.to_numeric(df_terrestre['Taxas'], errors='coerce')
        gerencial_terrestre['Valor total (R$)'] = (
            pd.to_numeric(df_terrestre['Total'], errors='coerce') + pd.to_numeric(df_terrestre['Taxas'], errors='coerce')
        )
        gerencial_terrestre['Tarifa Terrestre R$'] = gerencial_terrestre['Tarifa Terrestre R$'].fillna(0)
        gerencial_terrestre['Taxa Terrestre R$'] = gerencial_terrestre['Taxa Terrestre R$'].fillna(0)
        gerencial_terrestre['Valor total (R$)'] = gerencial_terrestre['Valor total (R$)'].fillna(0)
        #Função para criar a linha total no final do relatório
        total_terrestre = pd.DataFrame({
            'Diárias (Qted)': [gerencial_terrestre['Diárias (Qted)'].sum()],
            'Tarifa Terrestre R$': [gerencial_terrestre['Tarifa Terrestre R$'].sum()],
            'Valor total (R$)': [gerencial_terrestre['Valor total (R$)'].sum()]
        }, index=['Total'])
        gerencial_terrestre = pd.concat([gerencial_terrestre, total_terrestre])
    return gerencial_aereo, gerencial_terrestre
gerencial_aereo, gerencial_terrestre = gerar_bases_gerenciais(planilha_original)

# Exportar os arquivos em formato Excel com abas
excel_file = 'Relatorio_Gerencial_Final.xlsx'
with pd.ExcelWriter(excel_file) as writer:
    gerencial_aereo.to_excel(writer, sheet_name='Gerencial Aéreo', index=False)
    gerencial_terrestre.to_excel(writer, sheet_name='Gerencial Terrestre', index=False)
wb = load_workbook(excel_file)
#Estilização da Planilha Criada
for sheet in wb.sheetnames:
    ws = wb[sheet]
    # Formatação da primeira linha (títulos)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    # Formatação da última linha (total)
    for cell in ws[ws.max_row]:
        cell.fill = PatternFill(start_color='002060', end_color='002060', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True)
    # Aplicar texto vermelho para valores negativos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = Font(color='FF0000')  # Aplicar cor vermelha ao texto
wb.save(excel_file)
# Informar ao usuário onde o arquivo foi salvo
print(f"O arquivo '{excel_file}' foi salvo com sucesso.")
